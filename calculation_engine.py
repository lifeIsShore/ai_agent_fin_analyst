import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from models import InvoiceModel

def export_to_excel(invoice: InvoiceModel, output_filepath: str):
    """
    Exports the extracted invoice data to an Excel file with embedded formulas.
    This offloads the calculation verification to Excel, allowing non-tech users 
    to see exactly how totals are derived.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice Data"
    
    # Headers
    headers = ["Item Description", "Amount"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Write Items
    row_idx = 2
    for item in invoice.items:
        ws.cell(row=row_idx, column=1, value=item.description)
        # Store as float for math
        ws.cell(row=row_idx, column=2, value=float(item.amount)).number_format = '#,##0.00'
        row_idx += 1
        
    last_item_row = row_idx - 1
    
    # Write summary with formulas
    row_idx += 1
    
    # Subtotal
    ws.cell(row=row_idx, column=1, value="Calculated Subtotal").font = Font(bold=True)
    subtotal_cell = ws.cell(row=row_idx, column=2)
    # Excel Formula: SUM(B2:B{last_item_row})
    subtotal_cell.value = f"=SUM(B2:B{last_item_row})"
    subtotal_cell.number_format = '#,##0.00'
    subtotal_row = row_idx
    
    row_idx += 1
    
    # LLM Extracted Subtotal (For comparison)
    ws.cell(row=row_idx, column=1, value="Extracted Subtotal (LLM)").font = Font(italic=True)
    ws.cell(row=row_idx, column=2, value=float(invoice.subtotal)).number_format = '#,##0.00'
    
    row_idx += 1
    
    # Tax
    ws.cell(row=row_idx, column=1, value="Extracted Tax Amount").font = Font(italic=True)
    tax_cell = ws.cell(row=row_idx, column=2)
    tax_cell.value = float(invoice.tax_amount)
    tax_cell.number_format = '#,##0.00'
    tax_row = row_idx
    
    row_idx += 1
    
    # Final Total
    ws.cell(row=row_idx, column=1, value="Calculated Total").font = Font(bold=True)
    total_cell = ws.cell(row=row_idx, column=2)
    # Excel Formula: Subtotal + Tax
    total_cell.value = f"=B{subtotal_row}+B{tax_row}"
    total_cell.number_format = '#,##0.00'
    total_cell.font = Font(bold=True)
    
    row_idx += 1
    
    # Extracted Total
    ws.cell(row=row_idx, column=1, value="Extracted Total (LLM)").font = Font(italic=True)
    ws.cell(row=row_idx, column=2, value=float(invoice.total)).number_format = '#,##0.00'
    
    # Add metadata
    row_idx += 2
    ws.cell(row=row_idx, column=1, value="Metadata").font = Font(bold=True)
    ws.cell(row=row_idx+1, column=1, value="Invoice Number:")
    ws.cell(row=row_idx+1, column=2, value=invoice.invoice_number)
    ws.cell(row=row_idx+2, column=1, value="Date:")
    ws.cell(row=row_idx+2, column=2, value=invoice.date)
    ws.cell(row=row_idx+3, column=1, value="Bill To:")
    ws.cell(row=row_idx+3, column=2, value=invoice.bill_to)

    # Auto-adjust column width
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20

    wb.save(output_filepath)
    print(f"Excel report saved to {output_filepath}")
