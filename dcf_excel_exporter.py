import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List
from models_dcf import CompanyFinancials, DynamicScenarios

def export_dcf_to_excel(
    historical_data: List[CompanyFinancials], 
    dynamic_scenarios: DynamicScenarios,
    market_data: dict,
    wacc: float,
    scale: float,
    results: dict,
    output_filepath: str = "dcf_output.xlsx"
):
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: Valuation Results ---
    ws_val = wb.active
    ws_val.title = "Valuation Summary"
    
    ws_val.cell(row=1, column=1, value="DCF Valuation Summary").font = Font(bold=True, size=14)
    ws_val.cell(row=3, column=1, value="Ticker:").font = Font(bold=True)
    ws_val.cell(row=3, column=2, value=market_data.get('ticker', 'N/A'))
    ws_val.cell(row=4, column=1, value="WACC:").font = Font(bold=True)
    ws_val.cell(row=4, column=2, value=f"{wacc:.2%}")
    ws_val.cell(row=5, column=1, value="Assumed Scale Factor:").font = Font(bold=True)
    ws_val.cell(row=5, column=2, value=scale)
    
    row_idx = 7
    for scenario_name, data in results.items():
        ws_val.cell(row=row_idx, column=1, value=f"{scenario_name} Scenario").font = Font(bold=True, color="FFFFFF")
        ws_val.cell(row=row_idx, column=1).fill = PatternFill(start_color="4F81BD", fill_type="solid")
        
        ws_val.cell(row=row_idx+1, column=1, value="Assumed Revenue Growth")
        ws_val.cell(row=row_idx+1, column=2, value=f"{data['rev_growth']:.2%}")
        
        ws_val.cell(row=row_idx+2, column=1, value="Implied Share Price (DCF)")
        ws_val.cell(row=row_idx+2, column=2, value=data['implied_price_pg']).number_format = '#,##0.00'
        
        ws_val.cell(row=row_idx+3, column=1, value="Implied Share Price (Exit Multiple)")
        ws_val.cell(row=row_idx+3, column=2, value=data['implied_price_mult']).number_format = '#,##0.00'
        
        row_idx += 5
        
    ws_val.cell(row=row_idx, column=1, value="AI Insights Summary").font = Font(bold=True)
    ws_val.cell(row=row_idx+1, column=1, value=dynamic_scenarios.insight_summary)
    ws_val.column_dimensions['A'].width = 35
    ws_val.column_dimensions['B'].width = 20
    
    # --- Sheet 2: Historical Financials & KPIs ---
    ws_hist = wb.create_sheet("Historical Financials")
    headers = ["Metric"] + [f"Year {d.year}" for d in historical_data]
    for col_num, header in enumerate(headers, 1):
        cell = ws_hist.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4F81BD", fill_type="solid")
        
    metrics = [
        ("Revenue", lambda d: d.income_statement.revenue),
        ("EBIT", lambda d: d.income_statement.ebit),
        ("D&A", lambda d: d.income_statement.da),
        ("Net Income", lambda d: d.income_statement.net_income),
        ("Total Assets", lambda d: d.balance_sheet.total_assets),
        ("Total Debt", lambda d: d.balance_sheet.total_debt),
        ("Cash", lambda d: d.balance_sheet.cash),
        ("Operating Cash Flow", lambda d: d.cash_flow.operating_cash_flow),
        ("CapEx", lambda d: d.cash_flow.capex)
    ]
    
    for row_num, (metric_name, getter) in enumerate(metrics, 2):
        ws_hist.cell(row=row_num, column=1, value=metric_name).font = Font(bold=True)
        for col_num, data in enumerate(historical_data, 2):
            val = getter(data)
            ws_hist.cell(row=row_num, column=col_num, value=val).number_format = '#,##0'
            
    # Derived ML KPIs
    ws_hist.cell(row=13, column=1, value="--- ML Features (KPIs) ---").font = Font(bold=True)
    kpis = [
        ("EBIT Margin", lambda d: d.income_statement.ebit / d.income_statement.revenue if d.income_statement.revenue else 0),
        ("Net Margin", lambda d: d.income_statement.net_income / d.income_statement.revenue if d.income_statement.revenue else 0),
        ("CapEx / Revenue", lambda d: d.cash_flow.capex / d.income_statement.revenue if d.income_statement.revenue else 0),
        ("ROA (Net Income / Assets)", lambda d: d.income_statement.net_income / d.balance_sheet.total_assets if d.balance_sheet.total_assets else 0),
    ]
    for row_num, (metric_name, getter) in enumerate(kpis, 14):
        ws_hist.cell(row=row_num, column=1, value=metric_name).font = Font(italic=True)
        for col_num, data in enumerate(historical_data, 2):
            val = getter(data)
            ws_hist.cell(row=row_num, column=col_num, value=val).number_format = '0.00%'
            
    ws_hist.column_dimensions['A'].width = 30
    for i in range(2, len(headers) + 1):
        ws_hist.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 15

    # --- Sheet 3: Projections & ML Features (Base Case) ---
    ws_proj = wb.create_sheet("Projections (Base Case)")
    base_data = results['Base']
    proj_headers = ["Metric"] + [f"Projected Year {i}" for i in range(1, 6)]
    for col_num, header in enumerate(proj_headers, 1):
        cell = ws_proj.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="00B050", fill_type="solid")
        
    proj_metrics = [
        ("Revenue", base_data['projected_rev']),
        ("EBIT", base_data['projected_ebit']),
        ("D&A", base_data['projected_da']),
        ("CapEx", base_data['projected_capex']),
        ("Unlevered Free Cash Flow", base_data['projected_ufcf'])
    ]
    
    for row_num, (metric_name, values) in enumerate(proj_metrics, 2):
        ws_proj.cell(row=row_num, column=1, value=metric_name).font = Font(bold=True)
        for col_num, val in enumerate(values, 2):
            ws_proj.cell(row=row_num, column=col_num, value=val).number_format = '#,##0'
            
    # Projected KPIs
    ws_proj.cell(row=9, column=1, value="--- Projected ML Features ---").font = Font(bold=True)
    ws_proj.cell(row=10, column=1, value="EBIT Margin").font = Font(italic=True)
    ws_proj.cell(row=11, column=1, value="UFCF Margin").font = Font(italic=True)
    
    for col_num in range(2, 7):
        rev = base_data['projected_rev'][col_num-2]
        ebit = base_data['projected_ebit'][col_num-2]
        ufcf = base_data['projected_ufcf'][col_num-2]
        ws_proj.cell(row=10, column=col_num, value=(ebit/rev if rev else 0)).number_format = '0.00%'
        ws_proj.cell(row=11, column=col_num, value=(ufcf/rev if rev else 0)).number_format = '0.00%'
        
    ws_proj.column_dimensions['A'].width = 30
    for i in range(2, len(proj_headers) + 1):
        ws_proj.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 20

    # --- Sheet 4: Qualitative Insights ---
    ws_qual = wb.create_sheet("Qualitative Insights")
    ws_qual.cell(row=1, column=1, value="AI Qualitative Management Analysis").font = Font(bold=True, size=14)
    ws_qual.cell(row=3, column=1, value="Management Confidence Score (1-10):").font = Font(bold=True)
    ws_qual.cell(row=3, column=2, value=dynamic_scenarios.management_confidence_score).font = Font(bold=True, color="FF0000" if dynamic_scenarios.management_confidence_score < 5 else "00B050")
    ws_qual.cell(row=4, column=1, value="Confidence Rationale:").font = Font(bold=True)
    ws_qual.cell(row=4, column=2, value=dynamic_scenarios.confidence_rationale)
    ws_qual.cell(row=6, column=1, value="Overall Insight Summary:").font = Font(bold=True)
    ws_qual.cell(row=6, column=2, value=dynamic_scenarios.insight_summary)
    
    ws_qual.column_dimensions['A'].width = 35
    ws_qual.column_dimensions['B'].width = 80

    wb.save(output_filepath)
    print(f"Valuation results exported to {output_filepath}")
